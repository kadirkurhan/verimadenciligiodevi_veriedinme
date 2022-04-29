from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By

import pandas as pd
import time
from openpyxl import Workbook,load_workbook


wb = Workbook()

ws = wb.active
#wb = load_workbook(filename = path)

ws['A1'] = "MARKA"
ws['B1'] = "MODEL"
ws['C1'] = "BELLEK"
ws['D1'] = "ISLEMCI"
ws['E1'] = "ISLEMCIHIZI"
ws['F1'] = "DISK"
ws['G1'] = "EKRANBOYUTU"
ws['H1'] = "ISLETIMSISTEMI"
ws['I1'] = "EKRANCOZUNURLUGU"
ws['J1'] = "FIYAT"
ws['K1'] = "YORUM"

df = pd.read_excel (r'D:\Python\urunLinkleri.xlsx')

#df_product = pd.DataFrame(columns=['MARKA','MODEL','BELLEK','ISLEMCI','ISLEMCIHIZI','DISK','EKRANBOYUTU','ISLETIMSISTEMI','EKRANCOZUNURLUGU','YORUM'])
sayac=0
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
try:
    for i in df.values:
        driver.get(i[0])
        driver.maximize_window()
        try:
            driver.find_element(By.XPATH,'//*[@id="unf-prop"]/div/p/span').click()
            Fiyat = driver.find_element(By.XPATH,"/html/body/div[1]/div[3]/div/div[3]/div[1]/div[1]/div[2]/div[2]/div[1]/div[1]/div[1]/div/ins")
            Fiyat=Fiyat.get_attribute('content')
            Yorum = driver.find_element(By.XPATH,"//div/span/span")
            Yorum = int(Yorum.text)
            if(Yorum<5):
                Yorum=0
            elif(Yorum>=5):
                Yorum=1
            print(sayac)
            Marka = driver.find_element(By.XPATH,"//ul/li/p[text() ='Marka']/following-sibling::p[@class='unf-prop-list-prop']")
            Model = driver.find_element(By.XPATH,"//ul/li/p[text() ='Model']/following-sibling::p[@class='unf-prop-list-prop']")
            Bellek = driver.find_element(By.XPATH,"//ul/li/p[text() ='Bellek Kapasitesi']/following-sibling::p[@class='unf-prop-list-prop']")
            Islemci = driver.find_element(By.XPATH,"//ul/li/p[text() ='İşlemci']/following-sibling::p[@class='unf-prop-list-prop']")
            IslemciHizi = driver.find_element(By.XPATH,"//ul/li/p[text() ='İşlemci Hızı']/following-sibling::p[@class='unf-prop-list-prop']")
            DiskKapasitesi = driver.find_element(By.XPATH,"//ul/li/p[text() ='Disk Kapasitesi']/following-sibling::p[@class='unf-prop-list-prop']")
            EkranBoyutu = driver.find_element(By.XPATH,"//ul/li/p[text() ='Ekran Boyutu']/following-sibling::p[@class='unf-prop-list-prop']")
            IsletimSistemi = driver.find_element(By.XPATH,"//ul/li/p[text() ='İşletim Sistemi']/following-sibling::p[@class='unf-prop-list-prop']")
            EkranCozunurlugu = driver.find_element(By.XPATH,"//ul/li/p[text() ='Ekran Çözünürlüğü']/following-sibling::p[@class='unf-prop-list-prop']")
            ws.append([Marka.text,Model.text,Bellek.text,Islemci.text,IslemciHizi.text,DiskKapasitesi.text,EkranBoyutu.text,IsletimSistemi.text,EkranCozunurlugu.text,Fiyat,Yorum])
            sayac+=1
        except:
            continue
except:
    wb.save("D:/python/output.xlsx")
    #ws.append([Marka,Model,Bellek,Islemci.text,IslemciHizi.text,DiskKapasitesi.text,EkranBoyutu.text,IsletimSistemi.text,EkranCozunurlugu.text,Yorum.text])  # Sıradaki satıra sırasıyla dizi elemanlarını ekler
#df_product.append({'MARKA':Marka.text,'MODEL':Model.text,'BELLEK':Bellek.text,'ISLEMCI':Islemci.text,'ISLEMCIHIZI':IslemciHizi.text,'DISK':DiskKapasitesi.text,'EKRANBOYUTU':EkranBoyutu.text,'ISLETIMSISTEMI':İsletimSistemi.text,'EKRANCOZUNURLUGU':EkranCozunurlugu.text,'YORUM':Yorum.text},ignore_index=True)

wb.save("D:/python/output.xlsx")
